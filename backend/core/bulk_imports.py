# SPDX-License-Identifier: GPL-3.0-or-later
# Copyright (c) 2026 Slinky Software

"""Workbook generation and import helpers for bulk device and line onboarding."""

from __future__ import annotations

from io import BytesIO
import re
from typing import Any

from rest_framework import serializers

from provisioning.registry import get_device_type, list_device_types

from .models import Device, DeviceTypeConfig, Line, Site, mac_validator, normalize_mac
from .serializers import DeviceSerializer, LineSerializer


DEVICES_SHEET_NAME = "Devices"
LINES_SHEET_NAME = "Lines"
EXPECTED_SHEETS = [DEVICES_SHEET_NAME, LINES_SHEET_NAME]

DEVICE_HEADERS = [
    "MAC Address",
    "Name",
    "Model",
    "Site",
    "Line E164 Numbers",
    "Enabled",
]

LINE_HEADERS = [
    "Name",
    "Line E164 Number",
    "Registration User",
    "Registration Password",
    "Phone Label",
    "Shared",
]

MAX_IMPORT_ROWS = 2000
MAX_IMPORT_FILE_SIZE = 5 * 1024 * 1024
E164_PATTERN = re.compile(r"^\+?[0-9]{7,15}$")
TRUE_VALUES = {"1", "true", "t", "yes", "y"}
FALSE_VALUES = {"0", "false", "f", "no", "n"}
XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def generate_bulk_import_template() -> bytes:
    """Build the XLSX template used by the bulk import page."""
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    workbook = Workbook()
    devices_sheet = workbook.active
    devices_sheet.title = DEVICES_SHEET_NAME
    lines_sheet = workbook.create_sheet(LINES_SHEET_NAME)

    _configure_sheet(
        devices_sheet,
        DEVICE_HEADERS,
        {
            "A1": "Required. MAC address accepts colon, dash, or plain 12-hex input.",
            "B1": "Required. Friendly device name.",
            "C1": "Required. Exact backend device TypeID.",
            "D1": "Required. Existing Site name.",
            "E1": "Required. Comma-separated E164 numbers from the Lines sheet. First number becomes the primary line.",
            "F1": "Optional. TRUE or FALSE. Blank defaults to TRUE.",
        },
    )
    _configure_sheet(
        lines_sheet,
        LINE_HEADERS,
        {
            "A1": "Required. Friendly line name.",
            "B1": "Required. Unique line directory number for this import.",
            "C1": "Required. SIP registration username.",
            "D1": "Required. SIP registration password in plain text.",
            "E1": "Optional. Label shown on supported device screens.",
            "F1": "Optional. TRUE or FALSE. Blank defaults to FALSE.",
        },
    )

    type_ids = [device_type.TypeID for device_type in list_device_types()]
    if type_ids:
        type_list = ",".join(type_ids)
        if len(type_list) <= 255:
            device_type_validation = DataValidation(type="list", formula1=f'"{type_list}"', allow_blank=False)
            devices_sheet.add_data_validation(device_type_validation)
            device_type_validation.add(f"C2:C{MAX_IMPORT_ROWS + 1}")

    boolean_validation = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    devices_sheet.add_data_validation(boolean_validation)
    boolean_validation.add(f"F2:F{MAX_IMPORT_ROWS + 1}")

    shared_validation = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    lines_sheet.add_data_validation(shared_validation)
    shared_validation.add(f"F2:F{MAX_IMPORT_ROWS + 1}")

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def import_bulk_workbook(file_bytes: bytes) -> dict[str, Any]:
    """Validate and import workbook rows into line and device records."""
    from openpyxl import load_workbook

    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    _validate_workbook_structure(workbook)

    devices_sheet = workbook[DEVICES_SHEET_NAME]
    lines_sheet = workbook[LINES_SHEET_NAME]

    line_rows = _extract_rows(lines_sheet, LINE_HEADERS)
    device_rows = _extract_rows(devices_sheet, DEVICE_HEADERS)

    total_rows = len(line_rows) + len(device_rows)
    if total_rows > MAX_IMPORT_ROWS:
        raise serializers.ValidationError(
            {"detail": f"Workbook contains {total_rows} populated rows. Maximum supported rows is {MAX_IMPORT_ROWS}."}
        )

    summary = {
        "detail": "Bulk import completed.",
        "lines": {
            "total": len(line_rows),
            "imported_count": 0,
            "skipped_count": 0,
            "imported": [],
            "skipped": [],
        },
        "devices": {
            "total": len(device_rows),
            "imported_count": 0,
            "skipped_count": 0,
            "imported": [],
            "skipped": [],
        },
    }

    imported_lines_by_number, workbook_line_numbers = _import_lines(line_rows, summary["lines"])
    _import_devices(device_rows, summary["devices"], imported_lines_by_number, workbook_line_numbers)
    return summary


def validate_upload_file(uploaded_file) -> None:
    """Validate the uploaded file metadata before attempting to parse it."""
    if not uploaded_file:
        raise serializers.ValidationError({"detail": "An .xlsx workbook file is required."})

    if not uploaded_file.name.lower().endswith(".xlsx"):
        raise serializers.ValidationError({"detail": "Only .xlsx workbook files are supported."})

    if uploaded_file.size > MAX_IMPORT_FILE_SIZE:
        raise serializers.ValidationError(
            {"detail": f"Workbook is too large. Maximum supported file size is {MAX_IMPORT_FILE_SIZE // (1024 * 1024)} MB."}
        )


def _configure_sheet(sheet, headers: list[str], comments: dict[str, str]) -> None:
    """Apply shared formatting and documentation to a workbook sheet."""
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill

    sheet.append(headers)
    sheet.freeze_panes = "A2"

    header_fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[cell.column_letter].width = max(len(header) + 4, 18)

    for cell_ref, text in comments.items():
        sheet[cell_ref].comment = Comment(text, "PhoneManager")


def _validate_workbook_structure(workbook) -> None:
    """Ensure the uploaded workbook has the expected sheets and headers."""
    if workbook.sheetnames != EXPECTED_SHEETS:
        raise serializers.ValidationError(
            {"detail": "Workbook must contain exactly two sheets named Devices and Lines, in that order."}
        )


def _extract_rows(sheet, expected_headers: list[str]) -> list[dict[str, Any]]:
    """Return populated row dictionaries keyed by the expected headers."""
    actual_headers = [_cell_to_string(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    while actual_headers and actual_headers[-1] == "":
        actual_headers.pop()

    if actual_headers != expected_headers:
        raise serializers.ValidationError(
            {"detail": f"Sheet '{sheet.title}' headers do not match the expected template."}
        )

    rows: list[dict[str, Any]] = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = list(row)
        populated_values = [_cell_to_string(value) for value in values]
        if not any(populated_values[: len(expected_headers)]):
            continue

        if any(populated_values[len(expected_headers) :]):
            raise serializers.ValidationError(
                {"detail": f"Sheet '{sheet.title}' row {row_number} contains unexpected extra columns."}
            )

        row_data = {}
        for index, header in enumerate(expected_headers):
            value = values[index] if index < len(values) else None
            row_data[header] = value
        row_data["_row_number"] = row_number
        rows.append(row_data)

    return rows


def _import_lines(line_rows: list[dict[str, Any]], summary: dict[str, Any]) -> tuple[dict[str, Line], set[str]]:
    """Create all valid line rows and return them indexed by directory number."""
    existing_numbers = set(Line.objects.values_list("directory_number", flat=True))
    imported_lines_by_number: dict[str, Line] = {}
    workbook_numbers: set[str] = set()

    for row in line_rows:
        row_number = row["_row_number"]
        directory_number = _required_text(row, "Line E164 Number")
        identifier = directory_number or f"row {row_number}"

        if not directory_number:
            _record_skip(summary, row_number, identifier, "Line E164 Number is required.")
            continue

        if not E164_PATTERN.match(directory_number):
            _record_skip(summary, row_number, identifier, "Line E164 Number must be a valid E164 value.")
            continue

        if directory_number in workbook_numbers:
            _record_skip(summary, row_number, identifier, "Line E164 Number is duplicated within the workbook.")
            continue
        workbook_numbers.add(directory_number)

        if directory_number in existing_numbers:
            _record_skip(summary, row_number, identifier, "Line E164 Number already exists.")
            continue

        name = _required_text(row, "Name")
        registration_user = _required_text(row, "Registration User")
        registration_password = _required_text(row, "Registration Password")
        phone_label = _optional_text(row, "Phone Label")

        missing_fields = []
        if not name:
            missing_fields.append("Name")
        if not registration_user:
            missing_fields.append("Registration User")
        if not registration_password:
            missing_fields.append("Registration Password")
        if missing_fields:
            _record_skip(summary, row_number, identifier, f"Missing required fields: {', '.join(missing_fields)}.")
            continue

        if len(phone_label) > 128:
            _record_skip(summary, row_number, identifier, "Phone Label cannot exceed 128 characters.")
            continue

        try:
            is_shared = _parse_boolean(row.get("Shared"), default=False)
        except ValueError as exc:
            _record_skip(summary, row_number, identifier, str(exc))
            continue

        serializer = LineSerializer(
            data={
                "name": name,
                "directory_number": directory_number,
                "registration_account": registration_user,
                "registration_password": registration_password,
                "phone_label": phone_label,
                "is_shared": is_shared,
            }
        )
        if not serializer.is_valid():
            _record_skip(summary, row_number, identifier, _flatten_errors(serializer.errors))
            continue

        line = serializer.save()
        imported_lines_by_number[directory_number] = line
        summary["imported_count"] += 1
        summary["imported"].append(
            {
                "row": row_number,
                "id": line.id,
                "directory_number": line.directory_number,
                "name": line.name,
            }
        )

    summary["skipped_count"] = len(summary["skipped"])
    return imported_lines_by_number, workbook_numbers


def _import_devices(
    device_rows: list[dict[str, Any]],
    summary: dict[str, Any],
    imported_lines_by_number: dict[str, Line],
    workbook_line_numbers: set[str],
) -> None:
    """Create all valid devices after line creation has completed."""
    existing_macs = {normalize_mac(mac): True for mac in Device.objects.values_list("mac_address", flat=True)}
    sites = {site.name.strip().lower(): site for site in Site.objects.all()}
    device_defaults = {config.type_id: config.get_decrypted_device_defaults() for config in DeviceTypeConfig.objects.all()}
    workbook_macs: set[str] = set()

    for row in device_rows:
        row_number = row["_row_number"]
        mac_address_raw = _required_text(row, "MAC Address")
        identifier = mac_address_raw or f"row {row_number}"

        if not mac_address_raw:
            _record_skip(summary, row_number, identifier, "MAC Address is required.")
            continue

        normalized_mac = normalize_mac(mac_address_raw)
        try:
            mac_validator(normalized_mac)
        except Exception:
            _record_skip(summary, row_number, identifier, "MAC Address must contain 12 hexadecimal characters.")
            continue

        if normalized_mac in workbook_macs:
            _record_skip(summary, row_number, normalized_mac, "MAC Address is duplicated within the workbook.")
            continue
        workbook_macs.add(normalized_mac)

        if normalized_mac in existing_macs:
            _record_skip(summary, row_number, normalized_mac, "MAC Address already exists.")
            continue

        name = _required_text(row, "Name")
        site_name = _required_text(row, "Site")
        device_type_id = _required_text(row, "Model")
        line_numbers_raw = _required_text(row, "Line E164 Numbers")

        missing_fields = []
        if not name:
            missing_fields.append("Name")
        if not site_name:
            missing_fields.append("Site")
        if not device_type_id:
            missing_fields.append("Model")
        if not line_numbers_raw:
            missing_fields.append("Line E164 Numbers")
        if missing_fields:
            _record_skip(summary, row_number, normalized_mac, f"Missing required fields: {', '.join(missing_fields)}.")
            continue

        site = sites.get(site_name.lower())
        if not site:
            _record_skip(summary, row_number, normalized_mac, f"Site '{site_name}' was not found.")
            continue

        device_type = get_device_type(device_type_id)
        if not device_type:
            _record_skip(summary, row_number, normalized_mac, f"Model '{device_type_id}' is not a valid device TypeID.")
            continue

        try:
            enabled = _parse_boolean(row.get("Enabled"), default=True)
        except ValueError as exc:
            _record_skip(summary, row_number, normalized_mac, str(exc))
            continue

        line_numbers = [part.strip() for part in line_numbers_raw.split(",") if part and part.strip()]
        if not line_numbers:
            _record_skip(summary, row_number, normalized_mac, "At least one Line E164 Number is required.")
            continue

        if len(set(line_numbers)) != len(line_numbers):
            _record_skip(summary, row_number, normalized_mac, "Line E164 Numbers cannot contain duplicates within a device row.")
            continue

        invalid_numbers = [number for number in line_numbers if not E164_PATTERN.match(number)]
        if invalid_numbers:
            _record_skip(summary, row_number, normalized_mac, f"Invalid line numbers: {', '.join(invalid_numbers)}.")
            continue

        if len(line_numbers) > device_type.NumberOfLines:
            _record_skip(
                summary,
                row_number,
                normalized_mac,
                f"Model '{device_type_id}' supports at most {device_type.NumberOfLines} line(s).",
            )
            continue

        missing_from_workbook = [number for number in line_numbers if number not in workbook_line_numbers]
        if missing_from_workbook:
            _record_skip(
                summary,
                row_number,
                normalized_mac,
                "Referenced lines must exist on the Lines sheet: " + ", ".join(missing_from_workbook),
            )
            continue

        failed_line_rows = [number for number in line_numbers if number not in imported_lines_by_number]
        if failed_line_rows:
            _record_skip(
                summary,
                row_number,
                normalized_mac,
                "Referenced lines were not imported successfully: " + ", ".join(failed_line_rows),
            )
            continue

        ordered_lines = [imported_lines_by_number[number] for number in line_numbers]
        payload = {
            "name": name,
            "mac_address": normalized_mac,
            "device_type_id": device_type_id,
            "site": site.id,
            "line_1": ordered_lines[0].id,
            "lines": [line.id for line in ordered_lines[1:]],
            "device_specific_configuration": device_defaults.get(device_type_id, {}),
            "enabled": enabled,
        }

        serializer = DeviceSerializer(data=payload)
        if not serializer.is_valid():
            _record_skip(summary, row_number, normalized_mac, _flatten_errors(serializer.errors))
            continue

        device = serializer.save()
        extra_line_ids = [line.id for line in ordered_lines[1:]]
        if extra_line_ids:
            line_configuration = dict(device.line_configuration or {})
            line_configuration[Device.LINE_ORDER_KEY] = extra_line_ids
            device.line_configuration = line_configuration
            device.save(update_fields=["line_configuration"])

        summary["imported_count"] += 1
        summary["imported"].append(
            {
                "row": row_number,
                "id": device.id,
                "mac_address": device.mac_address,
                "name": device.description,
            }
        )

    summary["skipped_count"] = len(summary["skipped"])


def _required_text(row: dict[str, Any], key: str) -> str:
    """Return a trimmed string value for a required text cell."""
    return _optional_text(row, key)


def _optional_text(row: dict[str, Any], key: str) -> str:
    """Return a trimmed string value or an empty string for a cell."""
    value = row.get(key)
    if value is None:
        return ""
    return str(value).strip()


def _parse_boolean(value: Any, default: bool) -> bool:
    """Parse a spreadsheet boolean cell into a Python bool."""
    if value is None:
        return default
    if isinstance(value, bool):
        return value

    text = str(value).strip()
    if not text:
        return default

    lowered = text.lower()
    if lowered in TRUE_VALUES:
        return True
    if lowered in FALSE_VALUES:
        return False

    raise ValueError(f"Boolean fields must use TRUE or FALSE. Received '{text}'.")


def _cell_to_string(value: Any) -> str:
    """Normalize a workbook cell value into a comparable string."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value).strip()


def _flatten_errors(detail: Any) -> str:
    """Convert serializer error details into a readable single-line message."""
    if isinstance(detail, dict):
        parts = []
        for key, value in detail.items():
            parts.append(f"{key}: {_flatten_errors(value)}")
        return "; ".join(parts)

    if isinstance(detail, list):
        return "; ".join(_flatten_errors(item) for item in detail)

    return str(detail)


def _record_skip(summary: dict[str, Any], row_number: int, identifier: str, reason: str) -> None:
    """Append a skipped-row result entry."""
    summary["skipped"].append({
        "row": row_number,
        "identifier": identifier,
        "reason": reason,
    })
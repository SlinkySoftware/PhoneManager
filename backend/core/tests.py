# SPDX-License-Identifier: GPL-3.0-or-later
# Copyright (c) 2026 Slinky Software

"""Tests for bulk device and line import endpoints."""

from io import BytesIO

from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from django.utils.crypto import get_random_string
from openpyxl import Workbook, load_workbook
from rest_framework.test import APIClient

from .bulk_imports import DEVICE_HEADERS, DEVICES_SHEET_NAME, LINE_HEADERS, LINES_SHEET_NAME, XLSX_CONTENT_TYPE
from .models import Device, DeviceTypeConfig, Line, SIPServer, Site, UserProfile


class BulkImportApiTests(TestCase):
    """Exercise the XLSX template and import endpoints end to end."""

    def setUp(self):
        self.client = APIClient()

        self.admin_user = User.objects.create_user(username="admin", password="password")
        UserProfile.objects.create(
            user=self.admin_user,
            role=UserProfile.ROLE_ADMIN,
            auth_source=UserProfile.AUTH_SOURCE_LOCAL,
        )

        self.readonly_user = User.objects.create_user(username="viewer", password="password")
        UserProfile.objects.create(
            user=self.readonly_user,
            role=UserProfile.ROLE_READONLY,
            auth_source=UserProfile.AUTH_SOURCE_LOCAL,
        )

        self.sip_server = SIPServer.objects.create(name="Primary", host="pbx.example.test", port=5060)
        self.site = Site.objects.create(name="HQ", primary_sip_server=self.sip_server, timezone="UTC")

        self.existing_line = Line.objects.create(
            name="Existing",
            phone_label="Existing",
            directory_number="+19999999999",
            registration_account="existing-account",
            registration_password="existing-secret",
            is_shared=False,
        )
        Device.objects.create(
            mac_address="AA:BB:CC:DD:EE:FF",
            description="Existing Device",
            device_type_id="YealinkSIPT33G",
            site=self.site,
            line_1=self.existing_line,
            enabled=True,
        )

        self.device_type_config = DeviceTypeConfig.objects.create(type_id="YealinkSIPT33G")
        self.device_type_config.set_encrypted_device_defaults({"syslog_server": "10.0.0.5"})
        self.device_type_config.save()

    def test_download_bulk_import_template_has_expected_sheets_and_headers(self):
        self.client.force_authenticate(self.admin_user)

        response = self.client.get(reverse("bulk-import-template"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], XLSX_CONTENT_TYPE)

        workbook = load_workbook(filename=BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, [DEVICES_SHEET_NAME, LINES_SHEET_NAME])
        self.assertEqual(self._sheet_headers(workbook[DEVICES_SHEET_NAME]), DEVICE_HEADERS)
        self.assertEqual(self._sheet_headers(workbook[LINES_SHEET_NAME]), LINE_HEADERS)

    def test_bulk_import_rejects_invalid_workbook_structure(self):
        self.client.force_authenticate(self.admin_user)

        workbook = Workbook()
        workbook.active.title = "Wrong"
        workbook.create_sheet(LINES_SHEET_NAME)

        response = self.client.post(
            reverse("bulk-import-upload"),
            {"file": self._upload_file(workbook, "invalid.xlsx")},
            format="multipart",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("exactly two sheets named Devices and Lines", response.json()["detail"])

    def test_bulk_import_is_admin_only(self):
        self.client.force_authenticate(self.readonly_user)

        response = self.client.get(reverse("bulk-import-template"))

        self.assertEqual(response.status_code, 403)

    def test_bulk_import_creates_valid_rows_and_reports_skips(self):
        self.client.force_authenticate(self.admin_user)

        line_rows = [
            ["Desk 1 Primary", "+11111111111", "desk1-a", "secret-a", "Desk 1", "FALSE"],
            ["Desk 1 Secondary", "+12222222222", "desk1-b", "secret-b", "Desk 1B", "FALSE"],
            ["Desk 1 Tertiary", "+13333333333", "desk1-c", "secret-c", "Desk 1C", "TRUE"],
            ["Duplicate Existing", "+19999999999", "dup", "dup-secret", "Dup", "FALSE"],
        ]
        device_rows = [
            ["AA:BB:CC:DD:EE:01", "Desk 1", "YealinkSIPT33G", "HQ", "+11111111111,+13333333333,+12222222222", "TRUE"],
            ["AA:BB:CC:DD:EE:FF", "Duplicate MAC", "YealinkSIPT33G", "HQ", "+11111111111", "TRUE"],
            ["AA:BB:CC:DD:EE:02", "Failed Line Reference", "YealinkSIPT33G", "HQ", "+19999999999", "TRUE"],
        ]
        workbook = self._build_workbook(line_rows, device_rows)

        response = self.client.post(
            reverse("bulk-import-upload"),
            {"file": self._upload_file(workbook, "bulk-import.xlsx")},
            format="multipart",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["lines"]["total"], 4)
        self.assertEqual(payload["lines"]["imported_count"], 3)
        self.assertEqual(payload["lines"]["skipped_count"], 1)
        self.assertEqual(payload["devices"]["total"], 3)
        self.assertEqual(payload["devices"]["imported_count"], 1)
        self.assertEqual(payload["devices"]["skipped_count"], 2)

        imported_device = Device.objects.get(mac_address="AA:BB:CC:DD:EE:01")
        self.assertEqual(
            [line.directory_number for line in imported_device.get_ordered_lines()],
            ["+11111111111", "+13333333333", "+12222222222"],
        )
        retrieve_response = self.client.get(reverse("device-detail", args=[imported_device.id]))
        self.assertEqual(retrieve_response.status_code, 200)
        self.assertEqual(
            retrieve_response.json()["lines"],
            [line.id for line in imported_device.get_ordered_lines()[1:]],
        )
        self.assertEqual(imported_device.get_decrypted_device_config()["syslog_server"], "10.0.0.5")
        self.assertIn("already exists", payload["devices"]["skipped"][0]["reason"])
        self.assertIn("were not imported successfully", payload["devices"]["skipped"][1]["reason"])

    def _build_workbook(self, line_rows, device_rows) -> Workbook:
        workbook = Workbook()
        devices_sheet = workbook.active
        devices_sheet.title = DEVICES_SHEET_NAME
        lines_sheet = workbook.create_sheet(LINES_SHEET_NAME)

        devices_sheet.append(DEVICE_HEADERS)
        for row in device_rows:
            devices_sheet.append(row)

        lines_sheet.append(LINE_HEADERS)
        for row in line_rows:
            lines_sheet.append(row)

        return workbook

    def _sheet_headers(self, sheet):
        return [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    def _upload_file(self, workbook: Workbook, filename: str) -> SimpleUploadedFile:
        output = BytesIO()
        workbook.save(output)
        return SimpleUploadedFile(filename, output.getvalue(), content_type=XLSX_CONTENT_TYPE)


class UserApiTests(TestCase):
    """Exercise user management safety checks."""

    def setUp(self):
        self.client = APIClient()
        self.test_password = get_random_string(24)

        self.external_admin = User.objects.create_user(username="sso-admin", password=self.test_password)
        UserProfile.objects.create(
            user=self.external_admin,
            role=UserProfile.ROLE_ADMIN,
            auth_source=UserProfile.AUTH_SOURCE_SAML,
        )

    def test_delete_rejects_last_enabled_local_admin(self):
        target_user = User.objects.create_user(username="local-admin", password=self.test_password)
        UserProfile.objects.create(
            user=target_user,
            role=UserProfile.ROLE_ADMIN,
            auth_source=UserProfile.AUTH_SOURCE_LOCAL,
        )

        self.client.force_authenticate(self.external_admin)

        response = self.client.delete(reverse("user-detail", args=[target_user.id]))

        self.assertEqual(response.status_code, 409)
        self.assertEqual(response.json()["error_code"], "last_local_admin")
        self.assertTrue(User.objects.filter(pk=target_user.pk).exists())

    def test_delete_allows_local_admin_when_another_enabled_local_admin_exists(self):
        fallback_admin = User.objects.create_user(username="local-admin-2", password=self.test_password)
        UserProfile.objects.create(
            user=fallback_admin,
            role=UserProfile.ROLE_ADMIN,
            auth_source=UserProfile.AUTH_SOURCE_LOCAL,
        )
        target_user = User.objects.create_user(username="local-admin-1", password=self.test_password)
        UserProfile.objects.create(
            user=target_user,
            role=UserProfile.ROLE_ADMIN,
            auth_source=UserProfile.AUTH_SOURCE_LOCAL,
        )

        self.client.force_authenticate(self.external_admin)

        response = self.client.delete(reverse("user-detail", args=[target_user.id]))

        self.assertEqual(response.status_code, 204)
        self.assertFalse(User.objects.filter(pk=target_user.pk).exists())
        self.assertTrue(User.objects.filter(pk=fallback_admin.pk).exists())